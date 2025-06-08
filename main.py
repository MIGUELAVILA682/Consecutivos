from fastapi import FastAPI, Request, Form
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import pandas as pd
import openpyxl
import os
from datetime import datetime

app = FastAPI()

# estáticos y plantillas
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# Excel en OneDrive
RUTA_EXCEL = "consecutivos.xlsx"


# Usuario / contraseña prueba
USERS = {"admin": "1234", "usuario1": "clave1"}

# Sesión
user_sessions = {}

@app.get("/health")
async def health():
    return {"status": "ok"}

@app.get("/", response_class=HTMLResponse)
async def read_root(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

@app.post("/login")
async def login(request: Request, username: str = Form(...), password: str = Form(...)):
    if USERS.get(username) == password:
        user_sessions["logged_in_user"] = username
        return RedirectResponse("/dashboard", status_code=303)
    return templates.TemplateResponse("index.html",
        {"request": request, "error": "❌ Credenciales incorrectas"}, status_code=401)

@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request):
    if "logged_in_user" not in user_sessions:
        return RedirectResponse("/", status_code=303)

    hojas = []
    if os.path.exists(RUTA_EXCEL):
        hojas = pd.ExcelFile(RUTA_EXCEL).sheet_names

    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "hojas": hojas,
        "username": user_sessions["logged_in_user"]
    })

@app.post("/insert_project/")
async def insert_project(
    sheet_name: str = Form(...),
    asunto: str = Form(...),
    elaborado_por: str = Form(...)
):
    if "logged_in_user" not in user_sessions:
        return JSONResponse({"error": "No autenticado"}, status_code=401)

    try:
        # Abre libro y hoja
        wb = openpyxl.load_workbook(RUTA_EXCEL)
        ws = wb[sheet_name]

        # Busca primera fila vacía en Columna D (asunto)
        primera_vacía = None
        for row in range(2, ws.max_row + 2):
            if ws.cell(row=row, column=4).value is None:
                primera_vacía = row
                break

        if primera_vacía is None:
            return JSONResponse({"error": "No hay espacio disponible"}, status_code=400)

        # Calcula nuevo consecutivo a partir de la fila anterior
        anterior = ws.cell(row=primera_vacía-1, column=2).value
        if not anterior or not isinstance(anterior, str) or not anterior.startswith("C"):
            nuevo = "C000001"
        else:
            num = int(anterior[1:]) + 1
            nuevo = f"C{num:06d}"

        # Escribe usando .cell(...) para evitar MergedCellError
        ws.cell(row=primera_vacía, column=2).value = nuevo                              # B
        ws.cell(row=primera_vacía, column=3).value = datetime.now().strftime("%d/%m/%Y")  # C
        ws.cell(row=primera_vacía, column=4).value = asunto                               # D
        ws.cell(row=primera_vacía, column=5).value = elaborado_por                        # E

        wb.save(RUTA_EXCEL)
        wb.close()

        return JSONResponse({
            "message": f"✅ Agregado fila {primera_vacía} con consecutivo {nuevo}",
            "consecutivo": nuevo
        })

    except Exception as e:
        return JSONResponse({"error": f"Error interno: {e}"}, status_code=500)

@app.get("/logout")
async def logout():
    user_sessions.pop("logged_in_user", None)
    return RedirectResponse("/", status_code=303)
